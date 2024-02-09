from datetime import datetime

import openpyxl
import pandas as pd
from icecream import ic

from config.config import get_admin_id
from aiogram.types import CallbackQuery
from utils.bot import bot


def is_correct_time(data):
    if len(data) == 5 and data[2] in [':', '.', '/', '\\', ',', ';']:
        hour, minute = data[:2], data[3:5]

        if hour.isdigit() and 0 <= int(hour) <= 23 and minute.isdigit() and 0 <= int(minute) <= 59:
            return True
        return False
    return False


def refactor_datetime(time: str):
    colon_index = time.find(':')
    underlining_index = time.find('_')
    whitespace_index = time.find(' ')

    hour = time[colon_index - 2:colon_index]
    minute = time[colon_index + 1: colon_index + 3]
    if underlining_index != -1:
        type_of_mailing = time[underlining_index + 1:whitespace_index]
        return f"{hour}:{minute} - {type_of_mailing}"
    else:
        return f"{hour}:{minute}"


async def get_stub(call: CallbackQuery, l10n):
    await call.answer(l10n.format_value('in-development'))
    await call.answer()


async def send_message_to_admin(message, _reply_markup=None):
    await bot.send_message(get_admin_id(), message, reply_markup=_reply_markup)


def is_weekday():
    weekday = int(datetime.now().isoweekday())
    if 1 <= weekday <= 5:
        return True
    return False


def get_row_index(chat_id):
    df = pd.read_excel("data/main.xlsx")
    row_index = df.index[df['ID'] == chat_id].tolist()[0]
    return row_index


def save_data(chat_id, column_name, text):
    df = pd.read_excel("data/main.xlsx").astype("str")
    row_index = df.index[df['ID'] == str(chat_id)].tolist()[0]
    df.at[row_index, column_name] = text
    df.to_excel("data/main.xlsx", index=False)


headers = {
    "1.1": [3, 17, 8],
    "1.2": [21, 19, 9],
    "1.3": [41, 19, 9],

    "2.1": [3, 16, 7],
    "2.2": [20, 18, 9],
    "2.3": [39, 15, 7],

    "3.1": [3, 16, 8],
    "3.2": [20, 17, 9],

    "4.1": [3, 17, 8],
    "4.2": [21, 19, 10],
    "4.3": [41, 20, 10],
    "4.4": [62, 20, 10],
    "4.5": [83, 20, 10],
    "4.6": [104, 20, 10],
}


def get_question(header_nummer, nummer):
    workbook = openpyxl.load_workbook(f'data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    header_row = headers[header_nummer][0]
    return sheet.cell(row=header_row + nummer, column=2).value


def get_header(header_nummer):
    workbook = openpyxl.load_workbook(f'data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    return sheet.cell(row=headers[header_nummer][0], column=2).value


def save_answers(id_in_db, header_nummer, answers):
    workbook = openpyxl.load_workbook(f'data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    sheet.cell(row=headers[header_nummer][0], column=id_in_db).value = "-"

    for index, answer in enumerate(answers):
        sheet.cell(row=headers[header_nummer][0] + index + 1, column=id_in_db).value = answer

    workbook.save(f'data/section_{header_nummer[0]}.xlsx')


def get_status_section(id_in_db, header_nummer):
    workbook = openpyxl.load_workbook(f'data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    if sheet.cell(row=headers[header_nummer][0], column=id_in_db).value == "-":
        return "(Пройдено)"
    return " "


def get_id_in_db(chat_id):
    workbook = openpyxl.load_workbook('data/persons.xlsx')
    sheet = workbook['Лист1']

    for cell in sheet[1]:
        if str(cell.value) == str(chat_id):
            return cell.column


def is_exist(chat_id):
    workbook_persons = openpyxl.load_workbook('data/persons.xlsx')
    sheet = workbook_persons['Лист1']
    flag_id = -1
    for cell in sheet[1]:
        if str(chat_id) == str(cell.value):
            flag_id = cell.column

    return flag_id



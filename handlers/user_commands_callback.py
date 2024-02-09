import openpyxl
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery
from icecream import ic

from config.log_def import set_func, set_func_and_person
from handlers.questions_callback import menu_sections
from keyboards import inline
from handlers import auxiliary
from utils.bot import bot
from utils.states import UserState
from data.text import question_1_1, question_2_1
from config.config import structural_division

tag = "callback_handlers"


async def save_name_callback(call: CallbackQuery, state: FSMContext):
    func_name = "save_name_callback"
    set_func_and_person(func_name, tag, call.message)
    data = await state.get_data()
    name = data['name']

    workbook = openpyxl.load_workbook('data/persons.xlsx')
    sheet = workbook['Лист1']

    is_exist = -1
    for cell in sheet[1]:
        if call.message.chat.id == cell.value:
            is_exist = cell.column
    ic(is_exist)
    if is_exist == -1:
        sheet.cell(row=1, column=sheet.max_column+1).value = call.message.chat.id
        sheet.cell(row=2, column=sheet.max_column).value = name
        data['id_in_db'] = sheet.max_column

        for i in range(1, 5):
            workbook = openpyxl.load_workbook(f'data/section_{i}.xlsx')
            sheet = workbook['Лист1']
            sheet.cell(row=1, column=sheet.max_column + 1).value = call.message.chat.id
            sheet.cell(row=2, column=sheet.max_column).value = name
            workbook.save(f'data/section_{i}.xlsx')

    else:
        ic(name)
        sheet.cell(row=2, column=is_exist).value = name

        data['id_in_db'] = is_exist

    # data['header_nummer'] = 0
    await state.update_data(data)
    ic(data)

    workbook.save('data/persons.xlsx')
    await menu_sections(call, state)
    await call.answer()


async def rewrite_name_callback(call: CallbackQuery, state: FSMContext):
    func_name = "rewrite_name_callback"
    set_func_and_person(func_name, tag, call.message)

    await call.message.edit_text("Введите ваше ФИО через пробел.")
    await state.set_state(UserState.name)

#
# async def form_division_callback(call: CallbackQuery, state: FSMContext):
#     func_name = "form_division_callback"
#     set_func_and_person(func_name, tag, call.message)
#     await state.clear()
#     data = call.data[call.data.find('_')+1:]
#
#     is_montage = False
#     text = ""
#     match data:
#         case "admin":
#             text = "Администрация"
#         case "project":
#             text = "Проектный отдел"
#         case "service":
#             text = "Сервисный отдел"
#         case "montage":
#             text = "Отдел монтажа"
#             is_montage = True
#         case "cmto":
#             text = "СМТО"
#
#     auxiliary.save_data(call.message.chat.id, structural_division, text)
#
#     if is_montage:
#         await call.message.edit_text(question_2_1, reply_markup=inline.get_question_2_1())
#     else:
#         await call.message.edit_text(question_1_1, reply_markup=inline.get_question_1_1())




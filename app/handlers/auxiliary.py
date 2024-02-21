import openpyxl
import pandas as pd


def save_data(chat_id, column_name, text):
    df = pd.read_excel("app/data/main.xlsx").astype("str")
    row_index = df.index[df['ID'] == str(chat_id)].tolist()[0]
    df.at[row_index, column_name] = text
    df.to_excel("app/data/main.xlsx", index=False)


headers = {
    "1.1": [3, 17, 8],
    "1.2": [21, 19, 9],
    "1.3": [41, 19, 9],

    "2.1": [3, 16, 7],
    "2.2": [20, 18, 9],
    "2.3": [39, 15, 7],

    "3.1": [3, 16, 8],
    "3.2": [20, 17, 9],

    "4.1": [3, 17, 8],
    "4.2": [21, 19, 10],
    "4.3": [41, 20, 10],
    "4.4": [62, 20, 10],
    "4.5": [83, 20, 10],
    "4.6": [104, 20, 10],
}


def get_question(header_nummer, nummer):
    workbook = openpyxl.load_workbook(f'app/data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    header_row = headers[header_nummer][0]
    return sheet.cell(row=header_row + nummer, column=2).value


def get_header(header_nummer):
    workbook = openpyxl.load_workbook(f'app/data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    return sheet.cell(row=headers[header_nummer][0], column=2).value


def save_answers(id_in_db, header_nummer, answers):
    workbook = openpyxl.load_workbook(f'app/data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    sheet.cell(row=headers[header_nummer][0], column=id_in_db).value = "-"

    for index, answer in enumerate(answers):
        sheet.cell(row=headers[header_nummer][0] + index + 1, column=id_in_db).value = answer

    workbook.save(f'app/data/section_{header_nummer[0]}.xlsx')


def get_status_section(id_in_db, header_nummer):
    workbook = openpyxl.load_workbook(f'app/data/section_{header_nummer[0]}.xlsx')
    sheet = workbook['Лист1']

    if sheet.cell(row=headers[header_nummer][0], column=id_in_db).value == "-":
        return "(Пройдено)"
    return " "


def get_id_in_db(chat_id):
    workbook = openpyxl.load_workbook('app/data/persons.xlsx')
    sheet = workbook['Лист1']

    for cell in sheet[1]:
        if str(cell.value) == str(chat_id):
            return cell.column


def is_exist(chat_id):
    workbook_persons = openpyxl.load_workbook('app/data/persons.xlsx')
    sheet = workbook_persons['Лист1']
    flag_id = -1
    for cell in sheet[1]:
        if str(chat_id) == str(cell.value):
            flag_id = cell.column

    return flag_id



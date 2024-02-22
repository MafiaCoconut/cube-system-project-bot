import openpyxl
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery

from app.utils.logs import set_func_and_person
from app.handlers.questions_callback import menu_sections
from app.utils.states import UserState
from app.utils.postgresql import execute
tag = "callback_handlers"


async def save_name_callback(call: CallbackQuery, state: FSMContext):
    func_name = "save_name_callback"
    set_func_and_person(func_name, tag, call.message)
    data = await state.get_data()
    name = data['name']

    execute(f"INSERT INTO users (telegram_id, telegram_username) VALUES ('{call.message.chat.id}', '{name}')")

    # workbook_persons = openpyxl.load_workbook('app/data/persons.xlsx')
    # sheet = workbook_persons['Лист1']
    # is_exist = auxiliary.is_exist(call.message.chat.id)
    # if is_exist == -1:
    #     sheet.cell(row=1, column=sheet.max_column+1).value = call.message.chat.id
    #     sheet.cell(row=2, column=sheet.max_column).value = name
    #
    #     for i in range(1, 5):
    #         workbook = openpyxl.load_workbook(f'app/data/section_{i}.xlsx')
    #         sheet = workbook['Лист1']
    #         sheet.cell(row=1, column=sheet.max_column + 1).value = call.message.chat.id
    #         sheet.cell(row=2, column=sheet.max_column).value = name
    #         workbook.save(f'app/data/section_{i}.xlsx')
    #
    # else:
    #     sheet.cell(row=2, column=is_exist).value = name

    await state.update_data(data)
    await state.set_state()
    # workbook_persons.save('app/data/persons.xlsx')
    await menu_sections(call)
    await call.answer()


async def rewrite_name_callback(call: CallbackQuery, state: FSMContext):
    func_name = "rewrite_name_callback"
    set_func_and_person(func_name, tag, call.message)

    await call.message.edit_text("Введите ваше ФИО через пробел.")
    await state.set_state(UserState.name)
